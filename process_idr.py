import pandas as pd
from openpyxl import load_workbook

INPUT  = r'C:\Users\Zyvene\Downloads\lsdse-idr.xlsx'
OUTPUT = r'C:\Users\Zyvene\Downloads\lsdse-idr-summary.xlsx'
THRESHOLD = 0.015

drift_cols = ['drift_ratio_ex_x', 'drift_ratio_ex_y',
              'drift_ratio_ey_x', 'drift_ratio_ey_y']

df = pd.read_excel(INPUT)
df[drift_cols] = df[drift_cols].abs()
df = df[df[drift_cols].max(axis=1) > 0]

df['row_max'] = df[drift_cols].max(axis=1)

summary = (
    df.groupby('custom_building_id')
      .agg(
          Max_IDR    = ('row_max',     'max'),
          Num_Floors = ('story_level', 'max'),
      )
      .reset_index()
)

crit = (
    df.loc[df.groupby('custom_building_id')['row_max'].idxmax(),
           ['custom_building_id', 'story_level']]
      .rename(columns={'story_level': 'Critical_Story'})
)
summary = summary.merge(crit, on='custom_building_id')

summary['Status'] = summary['Max_IDR'].apply(
    lambda v: 'FAIL' if v > THRESHOLD else 'PASS'
)
summary['Structure'] = summary['custom_building_id'] + 1
summary = summary[['Structure', 'Num_Floors', 'Critical_Story', 'Max_IDR', 'Status']]

wb = load_workbook(INPUT)
if 'IDR Summary' in wb.sheetnames:
    del wb['IDR Summary']
ws = wb.create_sheet('IDR Summary', 0)

headers = ['Structure', 'No. of Floors', 'Critical Story',
           f'Max IDR (Threshold={THRESHOLD})', 'Status']
ws.append(headers)

for row in summary.itertuples(index=False):
    ws.append([row.Structure, row.Num_Floors, row.Critical_Story,
               row.Max_IDR, row.Status])

wb.save(OUTPUT)
print(f"Total : {len(summary)}")
print(f"PASS  : {(summary['Status']=='PASS').sum()}")
print(f"FAIL  : {(summary['Status']=='FAIL').sum()}")
print(f"Saved → {OUTPUT}")